from django.shortcuts import render, redirect
from .models import Expense
from .forms import ExpenseForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout

from django.shortcuts import get_object_or_404
from django.db.models import Sum
from django.db.models.functions import ExtractMonth
from django.http import HttpResponse
import openpyxl


from reportlab.platypus import SimpleDocTemplate
from reportlab.platypus import Table
from reportlab.platypus import TableStyle
from reportlab.platypus import Paragraph
from reportlab.platypus import Spacer

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from reportlab.platypus import PageBreak

from django.http import JsonResponse


from .forms import BudgetForm
from .models import Budget



from django.utils import timezone
from django.db.models import Sum
from django.db.models.functions import ExtractMonth, ExtractYear
from dateutil.relativedelta import relativedelta

from datetime import datetime

from datetime import timedelta

from .models import Category
from .forms import CategoryForm

from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font


def home(request):
    return render(request, 'home.html')

@login_required
def category_list(request):

    categories = Category.objects.filter(
        user=request.user
    )

    if request.method == 'POST':

        form = CategoryForm(
            request.POST
        )

        if form.is_valid():

            category = form.save(
                commit=False
            )

            category.user = request.user

            category.save()

            return redirect(
                'category_list'
            )

    else:

        form = CategoryForm()

    return render(
        request,
        'category_list.html',
        {
            'form': form,
            'categories': categories
        }
    )

@login_required
def delete_category(request, id):

    category = Category.objects.get(
        id=id,
        user=request.user
    )

    category.delete()

    return redirect(
        'category_list'
    )

@login_required
def dashboard(request):

    today = timezone.now().date()
    month_start = today.replace(day=1)

    # -----------------------------
    # EXPENSES
    # -----------------------------
    expenses = Expense.objects.filter(
        user=request.user
    ).select_related("category")

    categories = Category.objects.filter(user=request.user)

    search = request.GET.get("search")
    category_id = request.GET.get("category")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # ---------------- FILTER LOGIC ----------------
    if not start_date and not end_date:
        expenses = expenses.filter(
            expense_date__year=today.year,
            expense_date__month=today.month
        )
    elif start_date and end_date:
        expenses = expenses.filter(
            expense_date__range=[start_date, end_date]
        )

    if search:
        expenses = expenses.filter(title__icontains=search)

    if category_id:
        expenses = expenses.filter(category_id=category_id)

    expenses = expenses.order_by("-expense_date")

    # ---------------- TOTAL EXPENSE ----------------
    total = expenses.aggregate(
        total=Sum("amount")
    )["total"] or 0

    total_records = expenses.count()

    # ---------------- AI FORECAST ----------------
    amounts = [float(i.amount) for i in expenses]

    if amounts:
        predicted_expense, confidence = predict_expense(amounts)
    else:
        predicted_expense = 0
        confidence = 0

    # ---------------- BUDGET (FIXED ⭐) ----------------
    budget = Budget.objects.filter(
        user=request.user,
        month=month_start
    ).first()

    monthly_budget = 0
    remaining_amount = 0
    carry_forward = 0

    if budget:
        monthly_budget = budget.total_budget
        carry_forward = budget.carry_forward

        remaining_amount = budget.total_budget - total

        # sync DB safely
        if budget.spent != total or budget.remaining != remaining_amount:
            budget.spent = total
            budget.remaining = remaining_amount
            budget.save()

    # ---------------- CONTEXT ----------------
    context = {

        "expenses": expenses,
        "categories": categories,

        "total": total,
        "total_records": total_records,

        "predicted_expense": predicted_expense,
        "confidence": confidence,

        # BUDGET VALUES (SYNCED)
        "monthly_budget": monthly_budget,
        "remaining_amount": remaining_amount,
        "carry_forward": carry_forward,

        "current_month": today.strftime("%B"),

        "start_date": start_date,
        "end_date": end_date,
    }

    return render(request, "dashboard.html", context)
@login_required
def category_analytics(request):

    expenses = Expense.objects.filter(
        user=request.user
    ).select_related("category")

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if start_date and end_date:

        expenses = expenses.filter(
            expense_date__range=[
                start_date,
                end_date
            ]
        )

    categories = Category.objects.filter(
        user=request.user
    )

    category_totals = {}

    total_expense = 0

    for cat in categories:

        value = (

            expenses.filter(
                category=cat
            ).aggregate(
                Sum("amount")
            )["amount__sum"]

            or 0
        )

        category_totals[cat.name] = value

        total_expense += value

    context = {

        "category_totals":
        category_totals,

        "total_expense":
        total_expense,

        "categories":
        categories,

        "start_date":
        start_date,

        "end_date":
        end_date,
    }

    return render(
        request,
        "category_analytics.html",
        context
    )


from django.db.models import Sum
from datetime import datetime
from .models import Expense

@login_required
def expense_comparison(request):

    user_expenses = Expense.objects.filter(user=request.user)

    start = request.GET.get("start_date")
    end = request.GET.get("end_date")

    # DEFAULT: last 6 months
    if not start or not end:
        end_date = timezone.now().date()
        start_date = end_date - relativedelta(months=5)
    else:
        start_date = datetime.strptime(start, "%Y-%m-%d").date()
        end_date = datetime.strptime(end, "%Y-%m-%d").date()

    expenses = user_expenses.filter(
        expense_date__range=[start_date, end_date]
    )

    # GROUP BY MONTH + YEAR
    raw_data = expenses.annotate(
        month=ExtractMonth("expense_date"),
        year=ExtractYear("expense_date")
    ).values("month", "year").annotate(
        total=Sum("amount")
    )

    # MAP DATA
    data_map = {}
    for item in raw_data:
        key = f"{item['year']}-{item['month']}"
        data_map[key] = float(item["total"])

    # BUILD MONTH RANGE
    months = []
    totals = []

    temp_date = start_date.replace(day=1)

    while temp_date <= end_date:

        key = f"{temp_date.year}-{temp_date.month}"

        months.append(temp_date.strftime("%b %Y"))

        totals.append(round(data_map.get(key, 0), 2))

        temp_date += relativedelta(months=1)

    # STATS
    total_spent = sum(totals)
    avg_spent = total_spent / len(totals) if totals else 0
    max_month = months[totals.index(max(totals))] if totals else "N/A"

    growth = 0
    if len(totals) >= 2 and totals[-2] != 0:
        growth = ((totals[-1] - totals[-2]) / totals[-2]) * 100
    month_data = list(zip(months, totals))

    context = {
        "months": months,
        "totals": totals,
        "month_data": month_data,
        "start_date": start_date,
        "end_date": end_date,
        "total_spent": total_spent,
        "avg_spent": avg_spent,
        "max_month": max_month,
        "growth": growth,
    }

    return render(request, "expense_comparison.html", context)

@login_required
def add_expense(request):

    if request.method == 'POST':

        form = ExpenseForm(
            request.POST,
            user=request.user
        )

        if form.is_valid():

            expense = form.save(
                commit=False
            )

            expense.user = request.user

            expense.save()

            return JsonResponse({

                'success': True,

                'title': expense.title,

                'amount': str(expense.amount),

                'category': expense.category.name,

                'date': expense.expense_date.strftime('%Y-%m-%d')

            })

        return JsonResponse({

            'success': False,

            'errors': form.errors

        })

    form = ExpenseForm(
        user=request.user
    )

    return render(
        request,
        'add_expense.html',
        {
            'form': form
        }
    )
@login_required
def update_expense(request, id):

    expense = get_object_or_404(
        Expense,
        id=id,
        user=request.user
    )

    if request.method == 'POST':

        form = ExpenseForm(
        request.POST,
        instance=expense,
        user=request.user
        )

        if form.is_valid():

            form.save()

            return redirect('dashboard')

    else:

        form = ExpenseForm(
        instance=expense,
        user=request.user
        )

    return render(request,
                  'update_expense.html',
                  {'form': form})


@login_required
def delete_expense(request, id):

    expense = get_object_or_404(
        Expense,
        id=id,
        user=request.user
    )

    expense.delete()

    return redirect('dashboard')






@login_required
def export_excel(request):

    today = timezone.now().date()

    expenses = Expense.objects.filter(
        user=request.user
    ).select_related(
        "category"
    ).order_by(
        "-expense_date"
    )

    start_date = request.GET.get(
        "start_date"
    )

    end_date = request.GET.get(
        "end_date"
    )

    # FIX None
    if start_date == "None":
        start_date = ""

    if end_date == "None":
        end_date = ""

    # CURRENT MONTH
    if not start_date and not end_date:

        expenses = expenses.filter(
            expense_date__year=today.year,
            expense_date__month=today.month
        )

        report_period = (
            today.strftime(
                "%B %Y"
            )
        )

    # DATE RANGE
    elif start_date and end_date:

        expenses = expenses.filter(
            expense_date__range=[
                start_date,
                end_date
            ]
        )

        report_period = (
            f"{start_date} → {end_date}"
        )

    else:

        report_period = (
            today.strftime(
                "%B %Y"
            )
        )

    total = (
        expenses.aggregate(
            Sum("amount")
        )["amount__sum"]
        or 0
    )

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Expense Report"

    # HEADER
    sheet["A1"] = "Expense Report"
    sheet["A1"].font = Font(
        bold=True,
        size=18
    )

    sheet["A2"] = (
        f"Period: {report_period}"
    )

    # COLUMN TITLES
    headers = [
        "Title",
        "Amount",
        "Category",
        "Date"
    ]

    row = 4

    for col, value in enumerate(
        headers,
        1
    ):

        cell = sheet.cell(
            row=row,
            column=col
        )

        cell.value = value

        cell.font = Font(
            bold=True
        )

    # DATA
    row = 5

    for e in expenses:

        sheet.cell(
            row,
            1
        ).value = e.title

        sheet.cell(
            row,
            2
        ).value = float(
            e.amount
        )

        sheet.cell(
            row,
            3
        ).value = (
            e.category.name
        )

        sheet.cell(
            row,
            4
        ).value = (
            e.expense_date.strftime(
                "%d-%m-%Y"
            )
        )

        row += 1

    # TOTAL
    row += 2

    sheet.cell(
        row,
        1
    ).value = (
        "Total Expense"
    )

    sheet.cell(
        row,
        2
    ).value = float(
        total
    )

    sheet.cell(
        row,
        1
    ).font = Font(
        bold=True
    )

    sheet.cell(
        row,
        2
    ).font = Font(
        bold=True
    )

    # WIDTH
    sheet.column_dimensions[
        "A"
    ].width = 35

    sheet.column_dimensions[
        "B"
    ].width = 15

    sheet.column_dimensions[
        "C"
    ].width = 20

    sheet.column_dimensions[
        "D"
    ].width = 18

    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; filename="expense_report.xlsx"'
    )

    workbook.save(
        response
    )

    return response

@login_required
def export_pdf(request):

    today = timezone.now().date()

    expenses = Expense.objects.filter(
        user=request.user
    ).select_related(
        "category"
    )

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # DEFAULT → CURRENT MONTH
    if not start_date and not end_date:

        expenses = expenses.filter(
            expense_date__year=today.year,
            expense_date__month=today.month
        )

        report_title = today.strftime(
            "%B %Y"
        )

    # RANGE FILTER
    elif start_date and end_date:

        expenses = expenses.filter(
            expense_date__range=[
                start_date,
                end_date
            ]
        )

        report_title = (
            f"{start_date} → {end_date}"
        )

    else:

        report_title = "Expense Report"


    # TOTAL
    total_expense = sum(
        e.amount
        for e in expenses
    )


    response = HttpResponse(
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; filename="expense_report.pdf"'
    )

    p = canvas.Canvas(response)

    y = 800


    # HEADER
    p.setFont(
        "Helvetica-Bold",
        18
    )

    p.drawString(
        50,
        y,
        "Expense Report"
    )

    y -= 25


    p.setFont(
        "Helvetica",
        10
    )

    p.drawString(
        50,
        y,
        f"Period: {report_title}"
    )

    y -= 35


    # TABLE HEADER
    p.setFont(
        "Helvetica-Bold",
        10
    )

    p.drawString(50, y, "Date")
    p.drawString(140, y, "Title")
    p.drawString(310, y, "Category")
    p.drawString(460, y, "Amount")

    y -= 20

    p.line(
        50,
        y,
        550,
        y
    )

    y -= 20


    p.setFont(
        "Helvetica",
        10
    )

    for e in expenses:

        if y < 80:

            p.showPage()

            y = 800

        p.drawString(
            50,
            y,
            str(e.expense_date)
        )

        p.drawString(
            140,
            y,
            e.title[:25]
        )

        p.drawString(
            310,
            y,
            e.category.name
        )

        p.drawString(
            460,
            y,
            f"₹ {e.amount}"
        )

        y -= 20


    # ALWAYS SHOW TOTAL
    if y < 120:

        p.showPage()

        y = 800


    y -= 20

    p.line(
        50,
        y,
        550,
        y
    )

    y -= 30


    p.setFont(
        "Helvetica-Bold",
        14
    )

    p.drawString(
        50,
        y,
        f"Total Expense: ₹ {total_expense}"
    )

    p.save()

    return response
from datetime import date
import calendar





@login_required
def set_budget(request):

    today = date.today()

    # current month start
    month_start = date(today.year, today.month, 1)

    month_end = date(
        today.year,
        today.month,
        calendar.monthrange(today.year, today.month)[1]
    )

    # get current month budget
    budget = Budget.objects.filter(
        user=request.user,
        month=month_start
    ).first()

    # previous month carry forward
    previous = Budget.objects.filter(
        user=request.user
    ).exclude(
        month=month_start
    ).order_by("-month").first()

    carry = 0
    if previous:
        carry = max(previous.remaining, 0)

    # -------------------------
    # FORM SUBMIT
    # -------------------------
    if request.method == "POST":

        form = BudgetForm(request.POST, instance=budget)

        if form.is_valid():

            obj = form.save(commit=False)
            obj.user = request.user
            obj.month = month_start

            # carry forward
            obj.carry_forward = carry

            # total budget = new + carry
            obj.total_budget = (obj.budget_amount or 0) + carry

            obj.save()

            return redirect("set_budget")

    else:
        form = BudgetForm(instance=budget)

    # -------------------------
    # CURRENT MONTH SPENT
    # -------------------------
    spent = Expense.objects.filter(
        user=request.user,
        expense_date__year=today.year,
        expense_date__month=today.month
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0

    # -------------------------
    # BUDGET CALC
    # -------------------------
    current = 0
    remaining = 0
    percent = 0

    if budget:
        current = budget.total_budget
        remaining = current - spent

        budget.spent = spent
        budget.remaining = remaining
        budget.save()

        if current > 0:
            percent = round((spent / current) * 100)

    # -------------------------
    # HISTORY (LAST 12 MONTHS)
    # -------------------------
    history = Budget.objects.filter(
        user=request.user
    ).order_by("-month")[:12]

    # -------------------------
    # DAYS LEFT
    # -------------------------
    days_left = (month_end - today).days

    # -------------------------
    # INSIGHTS
    # -------------------------
    insights = []

    if percent > 80:
        insights.append("⚠ Budget usage above 80%")

    if remaining > 0:
        insights.append(f"💡 You can still save ₹{remaining}")

    # -------------------------
    # CONTEXT
    # -------------------------
    context = {

        "form": form,
        "current_budget": current,
        "spent": spent,
        "remaining": remaining,
        "percent": percent,
        "days_left": days_left,
        "carry": carry,
        "history": history,
        "month": today.strftime("%B %Y"),
        "insights": insights,

    }

    return render(request, "set_budget.html", context)



@login_required
def report(request):

    today = timezone.now().date()

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # ----------------------
    # EXPENSES
    # ----------------------
    expenses = Expense.objects.filter(
        user=request.user
    ).select_related("category")

    if start_date and end_date:
        expenses = expenses.filter(
            expense_date__range=[start_date, end_date]
        )
    else:
        expenses = expenses.filter(
            expense_date__year=today.year,
            expense_date__month=today.month
        )

    # ----------------------
    # TOTAL EXPENSE
    # ----------------------
    total = expenses.aggregate(
        total=Sum("amount")
    )["total"] or 0

    # ----------------------
    # BUDGET (FIXED ⭐ IMPORTANT)
    # ----------------------
    month_start = today.replace(day=1)

    budget = Budget.objects.filter(
        user=request.user,
        month=month_start
    ).first()

    total_budget = 0
    carry_forward = 0

    if budget:
        total_budget = budget.total_budget
        carry_forward = budget.carry_forward

    # ----------------------
    # BUDGET PERCENT
    # ----------------------
    budget_percent = 0

    if total_budget > 0:
        budget_percent = min(
            round((total / total_budget) * 100, 1),
            100
        )

    # ----------------------
    # CATEGORY REPORT
    # ----------------------
    category_report = expenses.values(
        "category__name"
    ).annotate(
        total=Sum("amount")
    ).order_by("-total")

    # ----------------------
    # TOP CATEGORY
    # ----------------------
    top_category = "-"
    top_amount = 0

    if category_report:
        top_category = category_report[0]["category__name"]
        top_amount = category_report[0]["total"]

    # ----------------------
    # RECENT + HIGHEST
    # ----------------------
    recent = expenses.order_by("-expense_date")[:5]
    highest = expenses.order_by("-amount").first()

    # ----------------------
    # DAILY AVG
    # ----------------------
    if start_date and end_date:
        days = (
            datetime.strptime(end_date, "%Y-%m-%d").date()
            - datetime.strptime(start_date, "%Y-%m-%d").date()
        ).days + 1
    else:
        days = today.day

    days = max(days, 1)
    avg_daily = round(total / days, 2)

    # ----------------------
    # INSIGHTS
    # ----------------------
    insights = []

    if total_budget > 0:
        if total > total_budget:
            insights.append("⚠ Budget exceeded this period")
        elif budget_percent >= 80:
            insights.append("⚠ Budget usage nearing limit")
        else:
            insights.append("✅ Budget spending looks healthy")

    if top_category != "-":
        insights.append(f"📊 Highest spending category: {top_category}")

    if highest:
        insights.append(f"💸 Largest expense ₹{highest.amount}")

    if total:
        insights.append(f"📅 Average daily spend ₹{avg_daily}")

    # ----------------------
    # CONTEXT
    # ----------------------
    context = {

        "total": total,

        # BUDGET VALUES (SYNCED)
        "monthly_budget": total_budget,
        "carry_forward": carry_forward,
        "total_budget": total_budget,

        "budget_percent": budget_percent,

        "top_category": top_category,
        "top_amount": top_amount,

        "recent": recent,
        "category_report": category_report,

        "insights": insights,

        "start_date": start_date,
        "end_date": end_date,

        "current_month": today.strftime("%B %Y"),

        "avg_daily": avg_daily,
        "highest": highest,
    }

    return render(request, "report.html", context)
        
@login_required
def export_category_excel(request):

    expenses = Expense.objects.filter(
        user=request.user
    ).select_related("category")

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if start_date and end_date:

        expenses = expenses.filter(
            expense_date__range=[
                start_date,
                end_date
            ]
        )

    categories = Category.objects.filter(
        user=request.user
    )

    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="category_analytics.xlsx"'
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "Category Analytics"

    ws.append([
        "Category",
        "Total Expense"
    ])

    total = 0

    for cat in categories:

        amount = (

            expenses

            .filter(
                category=cat
            )

            .aggregate(
                Sum("amount")
            )[
                "amount__sum"
            ]

            or 0
        )

        ws.append([
            cat.name,
            float(amount)
        ])

        total += amount

    ws.append([])

    ws.append([
        "TOTAL",
        float(total)
    ])

    wb.save(response)

    return response



@login_required
def export_category_pdf(request):

    expenses = Expense.objects.filter(
        user=request.user
    ).select_related("category")

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if start_date and end_date:

        expenses = expenses.filter(
            expense_date__range=[
                start_date,
                end_date
            ]
        )

    categories = Category.objects.filter(
        user=request.user
    )

    response = HttpResponse(
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="category_analytics.pdf"'
    )

    pdf = canvas.Canvas(response)

    y = 800

    pdf.setFont(
        "Helvetica-Bold",
        16
    )

    pdf.drawString(
        50,
        y,
        "Category Analytics"
    )

    y -= 40

    total = 0

    pdf.setFont(
        "Helvetica",
        11
    )

    for cat in categories:

        amount = (

            expenses

            .filter(
                category=cat
            )

            .aggregate(
                Sum("amount")
            )[
                "amount__sum"
            ]

            or 0
        )

        pdf.drawString(
            60,
            y,
            f"{cat.name}"
        )

        pdf.drawString(
            350,
            y,
            f"₹ {amount}"
        )

        total += amount

        y -= 25

        if y < 80:

            pdf.showPage()

            y = 800

    y -= 20

    pdf.setFont(
        "Helvetica-Bold",
        12
    )

    pdf.drawString(
        60,
        y,
        f"TOTAL : ₹ {total}"
    )

    pdf.save()

    return response